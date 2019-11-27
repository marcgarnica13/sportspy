import click
import os
import time
import logging
import logging.config

from openpyxl import load_workbook, Workbook

#Footballpy module
from model.footballpy.fs.loader import dfl
from model.match import Match

header_list = {
    1: [
        'centroid_pre_x',
        'centroid_pre_y',
        'dist_centroid_pre',
        'centroid_pre_a_x',
        'centroid_pre_a_y',
        'dist_centroid_pre_a',
        'centroid_pre_d_x',
        'centroid_pre_d_y',
        'dist_centroid_pre_d',
        'centroid_post_x',
        'centroid_post_y',
        'dist_centroid_post',
        'centroid_post_a_x',
        'centroid_post_a_y',
        'dist_centroid_post_a',
        'centroid_post_d_x',
        'centroid_post_d_y',
        'dist_centroid_post_d'
        ],
    2: [
        'dist_opp_ind_out',
        'dist_opp_ind_out_a',
        'dist_opp_ind_out_d',
        'dist_opp_ind_in',
        'dist_opp_ind_in_a',
        'dist_opp_ind_in_d'
        ],
    3: [
        'team_l_pre',
        'team_w_pre',
        'team_lw_pre',
        'team_l_pre_a',
        'team_w_pre_a',
        'team_lw_pre_a',
        'team_l_pre_d',
        'team_w_pre_d',
        'team_lw_pre_d',
        'team_l_post',
        'team_w_post',
        'team_lw_post',
        'team_l_post_a',
        'team_w_post_a',
        'team_lw_post_a',
        'team_l_post_d',
        'team_w_post_d',
        'team_lw_post_d',
        ],
    4: [
        'stretch_pre',
        'stretch_pre_a',
        'stretch_pre_d',
        'stretch_post',
        'stretch_post_a',
        'stretch_post_d',
        ],
    5: [
        'dist_centroid_out',
        'dist_centroid_out_a',
        'dist_centroid_out_d',
        'dist_centroid_in',
        'dist_centroid_in_a',
        'dist_centroid_in_d'
        ],
    6: [
        'dyadic_pre',
        'dyadic_pre_a',
        'dyadic_pre_d',
        'dyadic_post',
        'dyadic_post_a',
        'dyadic_post_d',
        ],
    7: [
        'dist_opp_pre',
        'dist_opp_pre_a',
        'dist_opp_pre_d',
        'dist_opp_post',
        'dist_opp_post_a',
        'dist_opp_post_d'
    ],
    8: [
        'pre_sc_home',
        'pre_sc_guest',
        'pre_sc_pitch_half1_h',
        'pre_sc_pitch_half1_a',
        'pre_sc_pitch_half2_h',
        'pre_sc_pitch_half2_a',
        'pre_sc_pitch_third1_h',
        'pre_sc_pitch_third1_a',
        'pre_sc_pitch_third2_h',
        'pre_sc_pitch_third2_a',
        'pre_sc_pitch_third3_h',
        'pre_sc_pitch_third3_a',
        'pre_sc_individual',

        'pre_a_sc_home',
        'pre_a_sc_guest',
        'pre_a_sc_pitch_half1_h',
        'pre_a_sc_pitch_half1_a',
        'pre_a_sc_pitch_half2_h',
        'pre_a_sc_pitch_half2_a',
        'pre_a_sc_pitch_third1_h',
        'pre_a_sc_pitch_third1_a',
        'pre_a_sc_pitch_third2_h',
        'pre_a_sc_pitch_third2_a',
        'pre_a_sc_pitch_third3_h',
        'pre_a_sc_pitch_third3_a',
        'pre_a_sc_individual',

        'pre_d_sc_home',
        'pre_d_sc_guest',
        'pre_d_sc_pitch_half1_h',
        'pre_d_sc_pitch_half1_a',
        'pre_d_sc_pitch_half2_h',
        'pre_d_sc_pitch_half2_a',
        'pre_d_sc_pitch_third1_h',
        'pre_d_sc_pitch_third1_a',
        'pre_d_sc_pitch_third2_h',
        'pre_d_sc_pitch_third2_a',
        'pre_d_sc_pitch_third3_h',
        'pre_d_sc_pitch_third3_a',
        'pre_d_sc_individual',

        'post_sc_home',
        'post_sc_guest',
        'post_sc_pitch_half1_h',
        'post_sc_pitch_half1_a',
        'post_sc_pitch_half2_h',
        'post_sc_pitch_half2_a',
        'post_sc_pitch_third1_h',
        'post_sc_pitch_third1_a',
        'post_sc_pitch_third2_h',
        'post_sc_pitch_third2_a',
        'post_sc_pitch_third3_h',
        'post_sc_pitch_third3_a',
        'post_sc_individual',

        'post_a_sc_home',
        'post_a_sc_guest',
        'post_a_sc_pitch_half1_h',
        'post_a_sc_pitch_half1_a',
        'post_a_sc_pitch_half2_h',
        'post_a_sc_pitch_half2_a',
        'post_a_sc_pitch_third1_h',
        'post_a_sc_pitch_third1_a',
        'post_a_sc_pitch_third2_h',
        'post_a_sc_pitch_third2_a',
        'post_a_sc_pitch_third3_h',
        'post_a_sc_pitch_third3_a',
        'post_a_sc_individual',

        'post_d_sc_home',
        'post_d_sc_guest',
        'post_d_sc_pitch_half1_h',
        'post_d_sc_pitch_half1_a',
        'post_d_sc_pitch_half2_h',
        'post_d_sc_pitch_half2_a',
        'post_d_sc_pitch_third1_h',
        'post_d_sc_pitch_third1_a',
        'post_d_sc_pitch_third2_h',
        'post_d_sc_pitch_third2_a',
        'post_d_sc_pitch_third3_h',
        'post_d_sc_pitch_third3_a',
        'post_d_sc_individual'
    ]
}

headers = {
    'Pdata': 'D',
    'MatchInformation': 'E',
    'Frame_S_pre': 'F',
    'Frame_E_pre': 'G',
    'Frame_S_post': 'H',
    'Frame_E_post': 'I',
    'Location': 'N',
    'PlayerOUT': 'U',
    'ID_OUT': 'U',
    'PlayerOUT': 'U',
    'PlayerOUT': 'U',
    'PlayerOUT': 'U',
}

def check_folder(dir_path, create=True):
    if not os.path.exists(dir_path):
        if create:
            os.makedirs(dir_path)
            return True
        else:
            return False
    else:
        return True

def add_debug_log(folder, logger):
    check_folder(folder)
    ch = logging.FileHandler(os.path.join(folder, 'debug.log'.format(folder)))
    ch.setLevel(logging.DEBUG)
    ch.setFormatter(logging.Formatter('%(asctime)s %(name)-12s %(levelname)-8s %(message)s'))
    logger.addHandler(ch)

variables = [
    'Team centroid',
    'Individual Distance to nearest opponent',
    'Team measures and ratio',
    'Stretch index',
    'Distance to centroid',
    'Dyadic distance',
    'Team distance to nearest opponent',
    'Space control'
]

@click.command()
@click.option('--test_file', '-t', type=click.Path(exists=True), required=True, help='Path to the excel test file commanding the test')
@click.option('--input_folder', '-i', type=click.Path(exists=True), required=True, help='Data folder path, match positions and information in the same folder')
@click.option('--output_file', '-o', type=click.Path(), help='Desired output folder for the test files')
@click.option('--variable', '-v', type=int, multiple=True, help='List of variables\n' + '\n'.join(['{}: {}'.format(i + 1, name) for i,name in enumerate(variables)]))
@click.option('--initial_row', '-irow', type=int, help='Initial row')
@click.option('--final_row', '-frow', type=int, help='Final row')
def run(test_file, input_folder, output_file, variable, initial_row, final_row):
    run_start_time = time.time()
    check_folder('logs')
    logging.config.fileConfig(fname=os.path.join('logging.ini'))
    logger = logging.getLogger()
    output_file = output_file or os.path.join('test')
    variable = variable or [i + 1 for i,name in enumerate(variables)]
    add_debug_log(output_file, logger)
    logger.info('Starting test with the following inputs')
    logger.info('- Input folder: {}'.format(input_folder))
    logger.info('- Test file coordinating the test: {}'.format(test_file))
    logger.info('- Test output folder: {}'.format(output_file))
    logger.info('- List of variables to compute: {}'.format(variable))
    m = Match(input_folder, logger)
    t_before = time.time()
    logger.debug('Match instance initialized in {} seconds'.format(time.time() - t_before))
    logger.info('Reading test controller file')
    t_before = time.time()
    controler = load_workbook(test_file).get_sheet_by_name('Subs')
    logger.debug('Test controlled file read in {} seconds'.format(time.time() - t_before))
    logger.info('Test file loaded')
    test_file_rows = (final_row or controler.max_row) - (initial_row or 0)
    test_file_columns = controler.max_column
    logger.debug('Test file contains {} rows and {} columns'.format(test_file_rows, test_file_columns))
    wb_idx = 1
    t_before = time.time()
    logger.info('Extracting header information')
    header = controler[wb_idx]
    header_dict = {}
    result_wb = Workbook()
    result_wb_sheet = result_wb.active
    result_wb_sheet.title = 'Run'
    for cell in header:
        header_dict[cell.value] = cell.column - 1
    wb_idx = initial_row or 2
    logger.debug('Header information extracted in {} seconds'.format(time.time() - t_before))
    result_row_count = 2
    while (wb_idx <= (final_row or controler.max_row)):
        row_start_time = time.time()
        logger.info('Processing row {} out of {}'.format(wb_idx, test_file_rows))
        logger_tag = "{} / {}".format(wb_idx, test_file_rows)
        row = controler[wb_idx]
        #Iniciar partido
        t_before = time.time()
        logger.info('Loading {} match information'.format(row[header_dict['Pdata']].value))
        load_result = m.load_match(logger_tag, row[header_dict['Pdata']].value, row[header_dict['MatchInformation']].value)
        if load_result:
            logger.info("Loading successfull after {} seconds".format(time.time() - t_before))
            #Set sub
            m.set_test(
                row[header_dict['Frame_S_pre']].value,
                row[header_dict['Frame_E_pre']].value,
                row[header_dict['Frame_S_post']].value,
                row[header_dict['Frame_E_post']].value,
                row[header_dict['Location']].value,
            )
            m.set_substitution(
                row[header_dict['ID_OUT']].value,
                row[header_dict['ID_IN']].value
            )
            #Compute vars
            t_before = time.time()
            try:
                test_results = m.run_test(variable)
                logger.info('Test results obtained in {} seconds'.format(time.time() - t_before))
                logger.debug(test_results)
                #Save results
                column_count = 1
                logger.info('Saving results to excel worksheet')
                logger.debug('Saving test data')
                logger.debug('### {}'.format(test_results))
                for key,item in header_dict.items():
                    result_wb_sheet.cell(row=result_row_count, column=item + 1).value = row[item].value
                    column_count += 1
                logger.debug('Added {} columns to the result from the test file'.format(column_count))
                for i, value in enumerate(test_results):
                    result_wb_sheet.cell(row=result_row_count, column=column_count).value = value
                    column_count += 1
                logger.debug('Finally adding {} columns to the result'.format(column_count))
            except Exception as e:
                logging.error('Error running this test:', exc_info=e)
            result_row_count += 1
        else:
            logger.info("Loading failed after {} seconds".format(time.time(), t_before))
        logger.debug('Row processed in {} seconds'.format(time.time() - row_start_time))
        logger.info('Row {} finished'.format(wb_idx))
        wb_idx += 1
    column_count = 1
    for key, item in header_dict.items():
        result_wb_sheet.cell(row=1, column=item + 1).value = key
        column_count += 1
    for i in variable:
        for header in header_list[i]:
            result_wb_sheet.cell(row=1, column=column_count).value = header
            column_count += 1
    result_wb.save(filename=os.path.join(output_file, 'test.xlsx').format(output_file))
    result_wb.close()
    logger.info('Test finished in {} seconds'.format(time.time() - run_start_time))




if __name__ == '__main__':
    run()
