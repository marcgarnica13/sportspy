from openpyxl import load_workbook, Workbook
import os

from model.match import Match

#%%
m = Match('/home/marc/Desktop/data/')

controler = load_workbook('/home/marc/Desktop/Bundes 16-17.xlsx').get_sheet_by_name('Subs')
wb_idx = 1
header = controler[wb_idx]
header_dict = {}
result_wb = Workbook()
result_wb_sheet = result_wb.active
result_wb_sheet.title = 'Run'
for cell in header:
    header_dict[cell.value] = cell.column - 1
print(header_dict)
#%%
wb_idx = 213

result_row_count = 1
row = controler[wb_idx]
print(row[header_dict['Pdata']].value)
#Iniciar partido
if m.load_match(row[header_dict['Pdata']].value, row[header_dict['MatchInformation']].value):
    print("Correct loading")

#%%
print(wb_idx)
#Set sub
m.set_test(
    row[header_dict['Frame_S_pre']].value,
    row[header_dict['Frame_E_pre']].value,
    row[header_dict['Frame_S_post']].value,
    row[header_dict['Frame_E_post']].value,
    row[header_dict['Location']].value,
)
#Compute vars
test_results = m.run_test()
print(test_results)
#Save results
column_count = 1
for key,item in header_dict.items():
    result_wb_sheet.cell(row=result_row_count, column=item)
    column_count += 1
for i, value in enumerate(test_results):
    result_wb_sheet.cell(row=result_row_count, column=column_count).value = value
    column_count += 1
result_row_count += 1

wb_idx += 1
result_wb.save(filename='test.xlsx')

